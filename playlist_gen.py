# splits, creates, & populates playlist for data collection
# creates forms for data collection
# 1-23-21

# == imports == #
import os
import spotipy 
import pandas as pd
from numpy import array_split
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment

# == main == #
# env variables
CLIENT_ID = os.environ.get('SPOTIFY_CLIENT_ID')
CLIENT_SECRET = os.environ.get('SPOTIFY_CLIENT_SECRET')
USERNAME = os.environ.get('SPOTIFY_USERNAME')

# spotify auth 
token = spotipy.util.prompt_for_user_token(
    username=USERNAME,
    scope='playlist-modify-private', 
    client_id=CLIENT_ID, 
    client_secret=CLIENT_SECRET, 
    redirect_uri='http://localhost:8888/callback' # registered with spotify
)

sp = spotipy.Spotify(auth=token)
user = sp.current_user()

# import playlist data
pl_path = 'classical_essentials'
pl_df = pd.read_csv(pl_path, index_col=0)

# create & populate playlists 
playlist_count = 5
playlists = array_split(pl_df,playlist_count)

for ix in range(playlist_count):
    playlist = sp.user_playlist_create(
        user = user['id'],
        name = f'ce_{ix+1}',
        public = False,
        collaborative=False,
        description=''
    )
    sp.user_playlist_add_tracks(
        user = user['id'],
        playlist_id = playlist['id'],
        tracks = playlists[i]['id'] # track ids
    )
    
    # create forms for survey
    playlists[ix].name = playlists[ix].name.apply(lambda x: x[:30]) # shorten names
    xlsx_path = f'forms/ce_{ix+1}.xlsx'
    playlists[ix][['name']].to_excel(xlsx_path, index=False)

    # create column for color input 
    wb = load_workbook(filename=xlsx_path)
    ws = wb.active
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 10
    b1 = ws['B1']
    b1.value = "Color" 
    b1.font = Font(bold=True)
    b1.alignment = Alignment(horizontal="center")
    wb.save(xlsx_path)

